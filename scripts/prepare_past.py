"""
prepare_past.py
令和5〜7年のExcelデータをJSONに変換するスクリプト

【使用データ（raw_data/に配置）】
  令和7年: 0932_*.xlsx (入所児童数), 0933_*.xlsx (受入可能数), 0934_*.xlsx (入所待ち人数)
  令和6年: 同様のファイル（年度別に命名が異なる場合がある）
  令和5年: 同様

【依存ライブラリ】
  pip install openpyxl

【実行方法】
  cd scripts/
  python3 prepare_past.py

【出力】
  ../data/r7_data.json
  ../data/r6_data.json
  ../data/r5_data.json
"""

import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("[ERROR] openpyxl が必要です: pip install openpyxl")
    sys.exit(1)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW_DIR = os.path.join(BASE_DIR, "raw_data")
DATA_DIR = os.path.join(BASE_DIR, "data")

AGE_COLS_JP = ["０歳", "１歳", "２歳", "３歳", "４歳", "５歳"]

# 年度別ファイル設定（横浜市オープンデータのファイル名に合わせて調整）
YEAR_CONFIG = {
    "R7": {
        "label": "令和7年",
        "output": "r7_data.json",
        "enrolled": "0932_",  # ファイル名プレフィックス
        "capacity": "0933_",
        "waiting":  "0934_",
    },
    # 令和6・5年のファイル番号が判明したら追加
    # "R6": {
    #     "label": "令和6年",
    #     "output": "r6_data.json",
    #     "enrolled": "XXXX_",
    #     "capacity": "YYYY_",
    #     "waiting":  "ZZZZ_",
    # },
}

def safe_int(val):
    try:
        v = str(val).strip().replace(",", "")
        if v in ("", "None", "-", "－", "―"):
            return 0
        return int(float(v))
    except (ValueError, TypeError):
        return 0

def find_file_by_prefix(directory, prefix):
    """ディレクトリからプレフィックスに一致するファイルを探す"""
    if not os.path.exists(directory):
        return None
    for fname in os.listdir(directory):
        if fname.startswith(prefix) and fname.endswith((".xlsx", ".xls")):
            return os.path.join(directory, fname)
    return None

def load_excel(filepath):
    """Excelファイルを読み込んで施設番号→年齢別数値の辞書を返す"""
    if not filepath or not os.path.exists(filepath):
        print(f"  [WARN] ファイルなし: {filepath}")
        return {}

    print(f"  [読込] {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # ヘッダー行を探す（施設番号・年齢列を含む行）
    header_row = None
    key_col_idx = None
    name_col_idx = None
    ward_col_idx = None
    age_col_indices = {}

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row is None:
            continue
        row_vals = [str(c).strip() if c is not None else "" for c in row]

        # 施設番号列を探す
        for i, val in enumerate(row_vals):
            if val in ("施設番号", "施設No", "施設ＮＯ", "番号"):
                key_col_idx = i
            if val in ("施設名称", "施設名", "保育所名"):
                name_col_idx = i
            if val in ("区", "区名"):
                ward_col_idx = i
            for age in AGE_COLS_JP:
                if val == age or val == age.replace("０","0").replace("１","1").replace("２","2").replace("３","3").replace("４","4").replace("５","5"):
                    age_col_indices[age] = i

        if key_col_idx is not None and len(age_col_indices) >= 3:
            header_row = row_idx
            break

    if header_row is None:
        print(f"  [WARN] ヘッダー行が見つかりません: {filepath}")
        return {}

    # データ行を読み込む
    data = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row is None or all(c is None for c in row):
            continue
        row_vals = [str(c).strip() if c is not None else "" for c in row]

        key = row_vals[key_col_idx] if key_col_idx is not None else ""
        if not key or key in ("施設番号", "合計", "計"):
            continue

        name = row_vals[name_col_idx] if name_col_idx is not None else ""
        ward = row_vals[ward_col_idx] if ward_col_idx is not None else ""

        age_data = {}
        for age in AGE_COLS_JP:
            idx = age_col_indices.get(age)
            age_data[age] = safe_int(row_vals[idx]) if idx is not None else 0

        data[key] = {
            "name": name,
            "ward": ward,
            "ages": age_data
        }

    print(f"    → {len(data)} 施設")
    return data

def process_year(config_key, config):
    print(f"\n=== {config['label']} データ変換 ===")

    enrolled_file = find_file_by_prefix(RAW_DIR, config["enrolled"])
    capacity_file = find_file_by_prefix(RAW_DIR, config["capacity"])
    waiting_file  = find_file_by_prefix(RAW_DIR, config["waiting"])

    enrolled_data = load_excel(enrolled_file)
    capacity_data = load_excel(capacity_file)
    waiting_data  = load_excel(waiting_file)

    if not enrolled_data:
        print(f"  [SKIP] {config['label']} の入所児童数データなし")
        return

    result = {}
    for key, info in enrolled_data.items():
        cap_info = capacity_data.get(key, {})
        wait_info = waiting_data.get(key, {})

        result[key] = {
            "id": key,
            "name": info.get("name", ""),
            "ward": info.get("ward", ""),
            "enrolled": info.get("ages", {}),
            "capacity": cap_info.get("ages", {a: 0 for a in AGE_COLS_JP}),
            "waiting":  wait_info.get("ages", {a: 0 for a in AGE_COLS_JP}),
        }

    output = {
        "year": config["label"],
        "label": config_key,
        "facilities": result
    }

    output_path = os.path.join(DATA_DIR, config["output"])
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"[完了] {output_path} に {len(result)} 施設のデータを出力")

def main():
    os.makedirs(DATA_DIR, exist_ok=True)

    if not os.path.exists(RAW_DIR):
        print(f"[ERROR] raw_data/ ディレクトリが見つかりません: {RAW_DIR}")
        print("       横浜市オープンデータのExcelファイルを raw_data/ に配置してください。")
        sys.exit(1)

    for config_key, config in YEAR_CONFIG.items():
        process_year(config_key, config)

    print("\n=== 全処理完了 ===")

if __name__ == "__main__":
    main()

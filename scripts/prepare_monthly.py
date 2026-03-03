"""
prepare_monthly.py
令和7年4月〜令和8年2月の月次データを生成するスクリプト

【Excelファイルの構造】
  0932_20260202.xlsx: 受入可能数（シート: R8.1, R7.12, ..., R7.4）
  0933_20260202.xlsx: 入所待ち人数
  0934_20260202.xlsx: 入所児童数

  令和8年2月のみCSV（既存のr8_202602.jsonを使用）

【出力先】
  ../data/monthly/r7_04.json 〜 r8_02.json（11ヶ月分）
"""

import csv
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("[ERROR] pip install openpyxl")
    sys.exit(1)

BASE_DIR    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW_DIR     = os.path.join(BASE_DIR, "raw_data")
MONTHLY_DIR = os.path.join(BASE_DIR, "data", "monthly")

AGE_COLS_CSV = ["０歳児", "１歳児", "２歳児", "３歳児", "４歳児", "５歳児"]
AGE_KEYS_APP = ["０歳",   "１歳",   "２歳",   "３歳",   "４歳",   "５歳"]

# シート名 → (年, 月, ラベル) のマッピング
SHEET_MAP = {
    "R7.4":  (7,  4,  "r7_04"),
    "R7.5":  (7,  5,  "r7_05"),
    "R7.6":  (7,  6,  "r7_06"),
    "R7.7":  (7,  7,  "r7_07"),
    "R7.8":  (7,  8,  "r7_08"),
    "R7.9":  (7,  9,  "r7_09"),
    "R7.10": (7,  10, "r7_10"),
    "R7.11": (7,  11, "r7_11"),
    "R7.12": (7,  12, "r7_12"),
    "R8.1":  (8,  1,  "r8_01"),
}

# ファイルマッピング（ファイル名 → データ種別）
# 実際の内容を確認した結果:
#   0932 = 受入可能数
#   0933 = 入所待ち人数
#   0934 = 入所児童数
FILE_ROLES = {
    "0932_20260202.xlsx": "capacity",
    "0933_20260202.xlsx": "waiting",
    "0934_20260202.xlsx": "enrolled",
}

def safe_int(val):
    try:
        v = str(val).strip().replace(",", "")
        if v in ("", "-", "－", "―", "None"):
            return 0
        return int(float(v))
    except (ValueError, TypeError):
        return 0

def load_sheet(wb, sheet_name):
    """シートからデータを読み込む。施設番号→{年齢: 数値}の辞書を返す"""
    if sheet_name not in wb.sheetnames:
        return {}

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # ヘッダー行を探す
    header_row_idx = None
    id_col = name_col = ward_col = None
    age_col_indices = {}

    for i, row in enumerate(rows):
        vals = [str(v).strip() if v is not None else "" for v in row]
        if "施設番号" in vals:
            header_row_idx = i
            id_col   = vals.index("施設番号")
            name_col = vals.index("施設・事業名") if "施設・事業名" in vals else None
            ward_col = vals.index("施設所在区") if "施設所在区" in vals else None
            for csv_col, app_key in zip(AGE_COLS_CSV, AGE_KEYS_APP):
                if csv_col in vals:
                    age_col_indices[app_key] = vals.index(csv_col)
            break

    if header_row_idx is None:
        return {}

    data = {}
    for row in rows[header_row_idx + 1:]:
        if row is None or all(v is None for v in row):
            continue
        vals = [str(v).strip() if v is not None else "" for v in row]

        facility_id = vals[id_col] if id_col < len(vals) else ""
        if not facility_id or facility_id in ("施設番号", "合計", ""):
            continue

        ages = {app_key: safe_int(vals[col_idx]) if col_idx < len(vals) else 0
                for app_key, col_idx in age_col_indices.items()}

        name = vals[name_col] if name_col is not None and name_col < len(vals) else ""
        ward = vals[ward_col] if ward_col is not None and ward_col < len(vals) else ""

        data[facility_id] = {"name": name, "ward": ward, "ages": ages}

    return data

def main():
    os.makedirs(MONTHLY_DIR, exist_ok=True)
    print("=== 月次データ変換開始 ===")

    # Excelファイル読み込み
    books = {}
    for fname, role in FILE_ROLES.items():
        path = os.path.join(RAW_DIR, fname)
        if not os.path.exists(path):
            print(f"[WARN] {fname} が見つかりません → raw_data/ にダウンロードしてください")
            continue
        print(f"[読込] {fname} ({role})")
        books[role] = openpyxl.load_workbook(path, data_only=True, read_only=True)

    if not books:
        print("[ERROR] Excelファイルが1つも読み込めませんでした")
        sys.exit(1)

    # 各月のデータを生成
    generated = []
    for sheet_name, (nengo, month, label) in SHEET_MAP.items():
        capacity_data = {}
        waiting_data  = {}
        enrolled_data = {}

        if "capacity" in books:
            capacity_data = load_sheet(books["capacity"], sheet_name)
        if "waiting" in books:
            waiting_data  = load_sheet(books["waiting"],  sheet_name)
        if "enrolled" in books:
            enrolled_data = load_sheet(books["enrolled"], sheet_name)

        if not enrolled_data and not capacity_data:
            print(f"  [SKIP] {sheet_name} データなし")
            continue

        # 施設IDの和集合
        all_ids = set(enrolled_data) | set(capacity_data) | set(waiting_data)

        facilities = {}
        for fid in all_ids:
            e = enrolled_data.get(fid, {})
            c = capacity_data.get(fid, {})
            w = waiting_data.get(fid, {})
            # 名前・区は最初に見つかったものを使用
            name = e.get("name") or c.get("name") or w.get("name") or ""
            ward = e.get("ward") or c.get("ward") or w.get("ward") or ""
            facilities[fid] = {
                "id":       fid,
                "name":     name,
                "ward":     ward,
                "enrolled": e.get("ages", {k: 0 for k in AGE_KEYS_APP}),
                "capacity": c.get("ages", {k: 0 for k in AGE_KEYS_APP}),
                "waiting":  w.get("ages", {k: 0 for k in AGE_KEYS_APP}),
            }

        output = {
            "year":         f"令和{nengo}年",
            "month":        f"{month}月",
            "label":        label,
            "displayLabel": f"令和{nengo}年{month}月",
            "facilities":   facilities,
        }

        out_path = os.path.join(MONTHLY_DIR, f"{label}.json")
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2)

        print(f"  [完了] {sheet_name} → {out_path} ({len(facilities)} 施設)")
        generated.append(label)

    # 令和8年2月のCSVからmonthly版を生成
    r8_csv_path = os.path.join(BASE_DIR, "data", "r8_202602.json")
    if os.path.exists(r8_csv_path):
        with open(r8_csv_path, encoding="utf-8") as f:
            r8 = json.load(f)
        r8["label"]        = "r8_02"
        r8["displayLabel"] = "令和8年2月"
        r8["month"]        = "2月"
        out_path = os.path.join(MONTHLY_DIR, "r8_02.json")
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(r8, f, ensure_ascii=False, indent=2)
        print(f"  [完了] R8.2 (CSV) → {out_path}")
        generated.append("r8_02")

    # for each book, close
    for book in books.values():
        book.close()

    print(f"\n=== 完了: {len(generated)} ヶ月分を生成 ===")
    print("生成ファイル:", generated)

if __name__ == "__main__":
    main()
